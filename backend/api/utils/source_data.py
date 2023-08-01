from io import BytesIO

from openpyxl import load_workbook


def read_employee_data(salary_file: BytesIO) -> dict:
    print('Reading employee sheet...')
    wb = load_workbook(salary_file)
    ws = wb['empleados']

    employee_data = {}
    for row in ws[f'A{ws.min_row + 1}:C{ws.max_row}']:
        # To know the existing employees
        employee_data[row[0].value] = {
            'id': row[0].value, 'name': row[1].value,
            'salary_base': float(row[2].value),
            'normal_week_day_mins': 0,
            'normal_week_day_value': 0,
            'normal_holiday_day_mins': 0,
            'normal_holiday_day_value': 0,
            'extra_week_day_mins': 0,
            'extra_week_day_value': 0,
            'extra_holiday_day_mins': 0,
            'extra_holiday_day_value': 0,
            'extra_week_night_mins': 0,
            'extra_week_night_value': 0,
            'extra_holiday_night_mins': 0,
            'extra_holiday_night_value': 0,
            'surcharge_week_mins': 0,
            'surcharge_week_value': 0,
            'surcharge_holiday_mins': 0,
            'surcharge_holiday_value': 0
        }

    return employee_data


def read_days_data(salary_file: BytesIO, employee_data: dict) -> tuple:
    print('Reading hours sheet...')
    wb = load_workbook(salary_file)
    ws = wb['horas']

    errors = []
    days_data = []
    row_num = 1
    for row in ws[f'A{ws.min_row + 1}:C{ws.max_row}']:
        row_num += 1
        if row[0].value not in employee_data.keys():
            errors.append(f'Fila {row_num}: La cédula no pertenece a ningún empleado')
            continue

        start_on = row[1].value
        end_on = row[2].value
        start_on = start_on.replace(second=0, microsecond=0)
        end_on = end_on.replace(second=0, microsecond=0)

        if start_on > end_on:
            errors.append(f'Fila {row_num}: La fecha de entrada es mayor a la de salida')
            continue

        date_diff = end_on - start_on
        hours_num = date_diff.total_seconds() / 60 / 60
        if hours_num > 24:
            errors.append(f'Fila {row_num}: La diferencia de la fecha de entrada y salida es mayor a 24 horas')
            continue

        days_data.append({
            'id': row[0].value, 'start_on': start_on, 'end_on': end_on
        })

    return days_data, errors

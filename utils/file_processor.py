from configparser import ConfigParser
from datetime import datetime, timedelta

from openpyxl import load_workbook, Workbook

config = ConfigParser()
config.read('utils/config.ini')
NORMAL_HOLIDAY = config.getfloat('normal', 'holiday')
SURCHARGE_WEEK = config.getfloat('surcharge', 'week')
SURCHARGE_HOLIDAY = config.getfloat('surcharge', 'holiday')
EXTRA_WEEK_DAY = config.getfloat('extra', 'week_day')
EXTRA_WEEK_NIGHT = config.getfloat('extra', 'week_night')
EXTRA_HOLIDAY_DAY = config.getfloat('extra', 'holiday_day')
EXTRA_HOLIDAY_NIGHT = config.getfloat('extra', 'holiday_night')


def read_employee_data() -> dict:
    print('Reading employee sheet...')
    wb = load_workbook(filename='utils/nomina.xlsx')
    ws = wb['empleados']

    employee_data = {'ids': []}
    for row in ws[f'A{ws.min_row + 1}:C{ws.max_row}']:
        # To know the existing employees
        employee_data['ids'].append(row[0].value)
        employee_data[row[0].value] = {
            'id': row[0].value, 'name': row[1].value,
            'salary_base': float(row[2].value),
            'normal_week_day_hours': 0,
            'normal_week_day_value': 0,
            'normal_holiday_day_hours': 0,
            'normal_holiday_day_value': 0,
            'extra_week_day_hours': 0,
            'extra_week_day_value': 0,
            'extra_holiday_day_hours': 0,
            'extra_holiday_day_value': 0,
            'extra_week_night_hours': 0,
            'extra_week_night_value': 0,
            'extra_holiday_night_hours': 0,
            'extra_holiday_night_value': 0,
            'surcharge_week_hours': 0,
            'surcharge_week_value': 0,
            'surcharge_holiday_hours': 0,
            'surcharge_holiday_value': 0
        }

    return employee_data


def read_days_data(employee_data: dict) -> tuple:
    print('Reading hours sheet...')
    wb = load_workbook(filename='utils/nomina.xlsx')
    ws = wb['horas']

    errors = ''
    days_data = []
    row_num = 1
    for row in ws[f'A{ws.min_row + 1}:D{ws.max_row}']:
        row_num += 1
        if row[1].value not in employee_data['ids']:
            errors += f'Fila {row_num}: La cédula no pertenece a ningún empleado'
            continue

        start_on = row[2].value
        end_on = row[3].value
        start_on = start_on.replace(second=0, microsecond=0)
        end_on = end_on.replace(second=0, microsecond=0)

        if start_on > end_on:
            errors += f'Fila {row_num}: La fecha de entrada es mayor a la de salida\n'
            continue

        date_diff = end_on - start_on
        hours_num = date_diff.total_seconds() / 60 / 60
        if hours_num > 24:
            errors += f'Fila {row_num}: La diferencia de la fecha de entra y salida es mayor a 24 horas\n'
            continue

        days_data.append({
            'id': row[1].value, 'start_on': start_on, 'end_on': end_on
        })

    return days_data, errors


def compute_salary_by_hours(days_data, employee_data, params):
    print('Calculando nomina...')
    for day in days_data:
        id = day['id']
        employee = employee_data.get(id)
        value_minute = employee['salary_base'] / 240 / 60
        start_on = day['start_on']
        end_on = day['end_on']
        num_minutes = 1
        while start_on < end_on:
            is_labor_day, is_noct_hour = get_type_hour(start_on)
            if num_minutes <= 480:
                if is_labor_day:
                    if is_noct_hour:
                        employee_data[id]['num_week_reacharge'] += 1 / 60
                        employee_data[id]['value_week_reacharge'] += (
                            value_minute * params['recargos_semana']
                        )
                    else:
                        employee_data[id]['num_diurnal_week'] += 1 / 60
                        employee_data[id]['value_diurnal_week'] += value_minute
                else:
                    if is_noct_hour:
                        employee_data[id]['num_noweek_reacharge'] += 1 / 60
                        employee_data[id]['value_noweek_reacharge'] += (
                            value_minute * params['recargos_dom_fest']
                        )
                    else:
                        employee_data[id]['num_diurnal_noweek'] += 1 / 60
                        employee_data[id]['value_diurnal_noweek'] += (
                            value_minute * params['valor_hora_dom_fest']
                        )
            else:
                if is_labor_day:
                    if is_noct_hour:
                        employee_data[id]['num_extra_noct_week'] += 1 / 60
                        employee_data[id]['value_extra_noct_week'] += (
                            value_minute * params['extra_noct_sem']
                        )
                    else:
                        employee_data[id]['num_extra_diurnal_week'] += 1 / 60
                        employee_data[id]['value_extra_diurnal_week'] += (
                            value_minute * params['extra_diurna_sem']
                        )
                else:
                    if is_noct_hour:
                        employee_data[id]['num_extra_noct_noweek'] += 1 / 60
                        employee_data[id]['value_extra_noct_noweek'] += (
                            value_minute * params['extra_noct_dom_fest']
                        )
                    else:
                        employee_data[id]['num_extra_diurnal_noweek'] += 1 / 60
                        employee_data[id]['value_extra_diurnal_noweek'] += (
                            value_minute * params['extra_diurna_dom_fest']
                        )
            start_on += timedelta(minutes=1)
            num_minutes += 1
        # print('Num minutes:', num_minutes-1)
        if num_minutes/60 < 1:
            print(end_on)

    descriptions = [
        {'Horas Ordinarias': ['num_diurnal_week', 'value_diurnal_week']},
        {'Horas Extras Diurnas': ['num_extra_diurnal_week', 'value_extra_diurnal_week']},
        {'Horas Extras Nocturna': ['num_extra_noct_week', 'value_extra_noct_week']},
        {'Recargos Nocturnos': ['num_week_reacharge', 'value_week_reacharge']},
        {'Horas Domin y Fest': ['num_diurnal_noweek', 'value_diurnal_noweek']},
        {'Horas Extras Diurnas Domin y Fest': ['num_extra_diurnal_noweek', 'value_extra_diurnal_noweek']},
        {'Horas Extras Nocturnas Domin y Fest': ['num_extra_noct_noweek', 'value_extra_noct_noweek']},
        {'Recargos Nocturnos Domin y Fest': ['num_noweek_reacharge', 'value_noweek_reacharge']}
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'salarios calculados'
    for key, employee in employee_data.items():
        ws.append(['CEDULA', '', key])
        ws.append(['NOMBRE', '', employee['name']])
        ws.append(['Descripción', 'Horas', 'Valor Horas'])
        for description in descriptions:
            for key_desc, values in description.items():
                if key_desc == 'Horas Ordinarias':
                    # print('Ordinarias', values)
                    hours = round(employee[values[0]]+16, 2)
                    hours_value = round(employee[values[1]]+(value_minute*960), 2)
                    row = [key_desc, hours, hours_value]
                    ws.append(row)
                elif employee[values[0]] != 0:
                    hours = round(employee[values[0]], 2)
                    hours_value = round(employee[values[1]], 2)
                    row = [key_desc, hours, hours_value]
                    ws.append(row)
        ws.append(['', '', ''])
    wb.save('./horas_procesadas.xlsx')
    print('Calculo de nomina hecha!')


def process_file():
    employee_data = read_employee_data()
    days_data, errors = read_days_data(employee_data)
    print(errors)
